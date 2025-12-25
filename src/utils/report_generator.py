"""
报告生成实用程序。

此模块提供：
- Excel 报告生成
- PDF 报告生成
- HTML 报告生成
- 报告模板
"""

import numpy as np
from typing import List, Dict, Optional, Union
from pathlib import Path
from datetime import datetime
import json

from src.logger import get_logger

logger = get_logger(__name__)


class ExcelReportGenerator:
    """
    生成带有统计数据和图表的 Excel 报告。
    
    使用 openpyxl 进行 Excel 文件生成。
    """
    
    def __init__(self):
        """初始化 ExcelReportGenerator。"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.chart import BarChart, PieChart, LineChart, Reference
            
            self.openpyxl = openpyxl
            self.Font = Font
            self.Alignment = Alignment
            self.PatternFill = PatternFill
            self.BarChart = BarChart
            self.PieChart = PieChart
            self.LineChart = LineChart
            self.Reference = Reference
            
        except ImportError:
            logger.error("未安装 openpyxl。请通过以下命令安装: pip install openpyxl")
            raise
    
    def generate_defect_report(self, statistics: Dict, output_path: str):
        """
        生成缺陷分析 Excel 报告。
        
        参数:
            statistics: 来自 compute_batch_statistics 的统计字典
            output_path: 输出 Excel 文件路径
        """
        logger.info(f"正在生成 Excel 报告: {output_path}")
        
        # 创建工作簿
        wb = self.openpyxl.Workbook()
        
        # 移除默认工作表
        wb.remove(wb.active)
        
        # 1. 摘要工作表
        self._create_summary_sheet(wb, statistics)
        
        # 2. 每张图像统计工作表
        self._create_per_image_sheet(wb, statistics)
        
        # 3. 尺寸分布工作表
        self._create_size_distribution_sheet(wb, statistics)
        
        # 保存工作簿
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        logger.info(f"Excel 报告已保存到: {output_path}")
    
    def _create_summary_sheet(self, wb, statistics: Dict):
        """创建摘要统计工作表。"""
        ws = wb.create_sheet("摘要")
        
        # 标题
        ws['A1'] = "缺陷分析摘要"
        ws['A1'].font = self.Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # 日期
        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:B2')
        
        # 统计数据
        row = 4
        header_fill = self.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = self.Font(color="FFFFFF", bold=True)
        
        # 表头
        ws[f'A{row}'] = "指标"
        ws[f'B{row}'] = "数值"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        
        row += 1
        
        # 数据
        metrics = [
            ("总图像数", statistics.get('total_images', 0)),
            ("已处理图像数", statistics.get('images_processed', 0)),
            ("有缺陷图像数", statistics.get('images_with_defects', 0)),
            ("无缺陷图像数", statistics.get('images_without_defects', 0)),
            ("总缺陷数", statistics.get('total_defects', 0)),
            ("平均每图缺陷数", f"{statistics.get('mean_defects_per_image', 0):.2f}"),
            ("平均覆盖率", f"{statistics.get('mean_coverage_ratio', 0):.4f}"),
            ("覆盖率标准差", f"{statistics.get('std_coverage_ratio', 0):.4f}"),
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_per_image_sheet(self, wb, statistics: Dict):
        """创建每张图像统计工作表。"""
        ws = wb.create_sheet("单图统计")
        
        # 表头
        headers = ["图像名称", "缺陷数量", "总面积", "覆盖率", 
                  "最大缺陷", "平均缺陷大小"]
        
        header_fill = self.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = self.Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = self.Alignment(horizontal='center')
        
        # 数据
        per_image_stats = statistics.get('per_image_stats', [])
        
        for row, stats in enumerate(per_image_stats, start=2):
            ws.cell(row=row, column=1, value=stats.get('image_name', ''))
            ws.cell(row=row, column=2, value=stats.get('num_defects', 0))
            ws.cell(row=row, column=3, value=stats.get('total_area', 0))
            ws.cell(row=row, column=4, value=f"{stats.get('coverage_ratio', 0):.4f}")
            ws.cell(row=row, column=5, value=stats.get('largest_defect', 0))
            ws.cell(row=row, column=6, value=f"{stats.get('mean_defect_size', 0):.2f}")
        
        # 调整列宽
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 18
    
    def _create_size_distribution_sheet(self, wb, statistics: Dict):
        """创建缺陷尺寸分布工作表。"""
        ws = wb.create_sheet("尺寸分布")
        
        # 表头
        ws['A1'] = "分箱范围"
        ws['B1'] = "频率"
        
        header_fill = self.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = self.Font(color="FFFFFF", bold=True)
        ws['A1'].fill = header_fill
        ws['B1'].fill = header_fill
        ws['A1'].font = header_font
        ws['B1'].font = header_font
        
        # 数据
        size_dist = statistics.get('defect_size_distribution', {})
        histogram = size_dist.get('histogram', [])
        bin_edges = size_dist.get('bin_edges', [])
        
        for i, (freq, bin_start) in enumerate(zip(histogram, bin_edges[:-1]), start=2):
            bin_end = bin_edges[i - 1]
            ws[f'A{i}'] = f"{bin_start:.0f} - {bin_end:.0f}"
            ws[f'B{i}'] = freq
        
        # 添加图表
        if histogram:
            chart = self.BarChart()
            chart.title = "缺陷尺寸分布"
            chart.x_axis.title = "尺寸范围 (像素)"
            chart.y_axis.title = "频率"
            
            data = self.Reference(ws, min_col=2, min_row=1, max_row=len(histogram) + 1)
            cats = self.Reference(ws, min_col=1, min_row=2, max_row=len(histogram) + 1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "D2")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def generate_training_report(self, history: Dict, output_path: str):
        """
        生成训练历史 Excel 报告。
        
        参数:
            history: 训练历史字典
            output_path: 输出 Excel 文件路径
        """
        logger.info(f"正在生成训练报告: {output_path}")
        
        wb = self.openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 创建指标工作表
        ws = wb.create_sheet("训练历史")
        
        # 表头
        headers = ["Epoch"] + list(history.keys())
        header_fill = self.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = self.Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # 数据
        num_epochs = len(history[list(history.keys())[0]])
        
        for epoch in range(num_epochs):
            ws.cell(row=epoch + 2, column=1, value=epoch + 1)
            
            for col, metric in enumerate(history.keys(), start=2):
                value = history[metric][epoch]
                ws.cell(row=epoch + 2, column=col, value=value)
        
        # 添加损失图表
        if 'train_loss' in history:
            chart = self.LineChart()
            chart.title = "损失曲线"
            chart.x_axis.title = "Epoch"
            chart.y_axis.title = "损失"
            
            train_loss_data = self.Reference(ws, min_col=2, min_row=1, max_row=num_epochs + 1)
            chart.add_data(train_loss_data, titles_from_data=True)
            
            if 'val_loss' in history:
                val_loss_col = list(history.keys()).index('val_loss') + 2
                val_loss_data = self.Reference(ws, min_col=val_loss_col, min_row=1, max_row=num_epochs + 1)
                chart.add_data(val_loss_data, titles_from_data=True)
            
            epochs_ref = self.Reference(ws, min_col=1, min_row=2, max_row=num_epochs + 1)
            chart.set_categories(epochs_ref)
            
            ws.add_chart(chart, f"A{num_epochs + 5}")
        
        # 保存
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        logger.info(f"训练报告已保存到: {output_path}")


class PDFReportGenerator:
    """
    生成带有 matplotlib 图形的 PDF 报告。
    
    使用 matplotlib 将图形保存到 PDF。
    """
    
    def __init__(self):
        """初始化 PDFReportGenerator。"""
        try:
            from matplotlib.backends.backend_pdf import PdfPages
            self.PdfPages = PdfPages
        except ImportError:
            logger.error("matplotlib 未正确安装")
            raise
    
    def generate_defect_report(self, statistics: Dict, figures: List,
                              output_path: str):
        """
        生成缺陷分析 PDF 报告。
        
        参数:
            statistics: 统计字典
            figures: matplotlib 图形列表
            output_path: 输出 PDF 文件路径
        """
        logger.info(f"正在生成 PDF 报告: {output_path}")
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with self.PdfPages(output_path) as pdf:
            # 添加标题页
            self._add_title_page(pdf, statistics)
            
            # 添加所有图形
            for fig in figures:
                pdf.savefig(fig, bbox_inches='tight')
            
            # 添加元数据
            d = pdf.infodict()
            d['Title'] = '缺陷分析报告'
            d['Author'] = '工业缺陷分割系统'
            d['Subject'] = '缺陷统计与可视化'
            d['CreationDate'] = datetime.now()
        
        logger.info(f"PDF 报告已保存到: {output_path}")
    
    def _add_title_page(self, pdf, statistics: Dict):
        """向 PDF 添加标题页。"""
        import matplotlib.pyplot as plt
        
        fig = plt.figure(figsize=(8.5, 11))
        fig.text(0.5, 0.7, '缺陷分析报告', 
                ha='center', fontsize=24, fontweight='bold')
        fig.text(0.5, 0.6, f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                ha='center', fontsize=12)
        
        # 添加摘要统计
        summary_text = f"""
        总图像数: {statistics.get('total_images', 0)}
        有缺陷图像数: {statistics.get('images_with_defects', 0)}
        总缺陷数: {statistics.get('total_defects', 0)}
        平均每图缺陷数: {statistics.get('mean_defects_per_image', 0):.2f}
        平均覆盖率: {statistics.get('mean_coverage_ratio', 0):.4f}
        """
        
        fig.text(0.5, 0.4, summary_text, ha='center', fontsize=11,
                family='monospace')
        
        plt.axis('off')
        pdf.savefig(fig, bbox_inches='tight')
        plt.close(fig)


class HTMLReportGenerator:
    """
    生成带有嵌入图表的 HTML 报告。
    
    创建包含统计数据和可视化的独立 HTML 文件。
    """
    
    def __init__(self):
        """初始化 HTMLReportGenerator。"""
        pass
    
    def generate_defect_report(self, statistics: Dict, chart_paths: Dict[str, str],
                              output_path: str):
        """
        生成缺陷分析 HTML 报告。
        
        参数:
            statistics: 统计字典
            chart_paths: 将图表名称映射到文件路径的字典
            output_path: 输出 HTML 文件路径
        """
        logger.info(f"正在生成 HTML 报告: {output_path}")
        
        html_content = self._create_html_template(statistics, chart_paths)
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"HTML 报告已保存到: {output_path}")
    
    def _create_html_template(self, statistics: Dict, chart_paths: Dict[str, str]) -> str:
        """创建 HTML 报告模板。"""
        
        # 页眉
        html = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>缺陷分析报告</title>
    <style>
        body {{
            font-family: "Microsoft YaHei", Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #34495e;
            margin-top: 30px;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .stat-card {{
            background-color: #ecf0f1;
            padding: 20px;
            border-radius: 8px;
            border-left: 4px solid #3498db;
        }}
        .stat-label {{
            font-size: 14px;
            color: #7f8c8d;
            margin-bottom: 5px;
        }}
        .stat-value {{
            font-size: 28px;
            font-weight: bold;
            color: #2c3e50;
        }}
        .chart-container {{
            margin: 30px 0;
            text-align: center;
        }}
        .chart-container img {{
            max-width: 100%;
            height: auto;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        .footer {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #bdc3c7;
            text-align: center;
            color: #7f8c8d;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🔍 缺陷分析报告</h1>
        <p><strong>生成时间:</strong> {timestamp}</p>
        
        <h2>📊 摘要统计</h2>
        <div class="stats-grid">
            <div class="stat-card">
                <div class="stat-label">总图像数</div>
                <div class="stat-value">{total_images}</div>
            </div>
            <div class="stat-card">
                <div class="stat-label">有缺陷图像数</div>
                <div class="stat-value">{images_with_defects}</div>
            </div>
            <div class="stat-card">
                <div class="stat-label">总缺陷数</div>
                <div class="stat-value">{total_defects}</div>
            </div>
            <div class="stat-card">
                <div class="stat-label">平均每图缺陷数</div>
                <div class="stat-value">{mean_defects:.2f}</div>
            </div>
            <div class="stat-card">
                <div class="stat-label">平均覆盖率</div>
                <div class="stat-value">{mean_coverage:.4f}</div>
            </div>
        </div>
""".format(
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            total_images=statistics.get('total_images', 0),
            images_with_defects=statistics.get('images_with_defects', 0),
            total_defects=statistics.get('total_defects', 0),
            mean_defects=statistics.get('mean_defects_per_image', 0),
            mean_coverage=statistics.get('mean_coverage_ratio', 0)
        )
        
        # 添加图表
        html += "\n        <h2>📈 可视化</h2>\n"
        
        for chart_name, chart_path in chart_paths.items():
            # 使用相对路径
            rel_path = Path(chart_path).name
            html += f"""
        <div class="chart-container">
            <h3>{chart_name.replace('_', ' ').title()}</h3>
            <img src="{rel_path}" alt="{chart_name}">
        </div>
"""
        
        # 页脚
        html += """
        <div class="footer">
            <p>由工业缺陷分割系统生成</p>
        </div>
    </div>
</body>
</html>
"""
        
        return html


class ReportManager:
    """
    管理报告生成工作流程。
    
    协调统计计算、可视化和报告生成。
    """
    
    def __init__(self):
        """初始化 ReportManager。"""
        self.excel_generator = None
        self.pdf_generator = None
        self.html_generator = HTMLReportGenerator()
    
    def generate_complete_report(self, mask_paths: List[str],
                                output_dir: str,
                                report_formats: List[str] = ['excel', 'pdf', 'html']):
        """
        以多种格式生成完整报告。
        
        参数:
            mask_paths: 掩码文件路径列表
            output_dir: 报告输出目录
            report_formats: 要生成的格式列表 ('excel', 'pdf', 'html')
        """
        from src.utils.statistics import DefectStatistics
        from src.utils.visualization import DefectVisualizer
        
        logger.info("正在开始生成完整报告...")
        
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # 1. 计算统计数据
        logger.info("正在计算统计数据...")
        defect_stats = DefectStatistics()
        statistics = defect_stats.compute_batch_statistics(mask_paths)
        
        # 保存统计数据 JSON
        stats_path = output_dir / 'statistics.json'
        with open(stats_path, 'w') as f:
            json.dump(statistics, f, indent=2)
        logger.info(f"统计数据已保存到: {stats_path}")
        
        # 2. 生成可视化
        logger.info("正在生成可视化...")
        visualizer = DefectVisualizer()
        
        figures = {}
        chart_paths = {}
        
        # 缺陷尺寸分布
        all_defect_areas = []
        for img_stats in statistics.get('per_image_stats', []):
            all_defect_areas.extend(img_stats.get('defect_areas', []))
        
        if all_defect_areas:
            fig1 = visualizer.plot_defect_size_distribution(
                all_defect_areas,
                output_path=str(output_dir / 'defect_size_distribution.png')
            )
            figures['size_distribution'] = fig1
            chart_paths['size_distribution'] = str(output_dir / 'defect_size_distribution.png')
        
        # 每张图像的缺陷计数
        defect_counts = [stats['num_defects'] for stats in statistics.get('per_image_stats', [])]
        if defect_counts:
            fig2 = visualizer.plot_defect_count_per_image(
                defect_counts,
                output_path=str(output_dir / 'defect_count_per_image.png')
            )
            figures['defect_counts'] = fig2
            chart_paths['defect_counts'] = str(output_dir / 'defect_count_per_image.png')
        
        # 覆盖率分布
        coverage_ratios = [stats['coverage_ratio'] for stats in statistics.get('per_image_stats', [])
                          if stats['coverage_ratio'] > 0]
        if coverage_ratios:
            fig3 = visualizer.plot_coverage_ratio_distribution(
                coverage_ratios,
                output_path=str(output_dir / 'coverage_ratio_distribution.png')
            )
            figures['coverage_ratio'] = fig3
            chart_paths['coverage_ratio'] = str(output_dir / 'coverage_ratio_distribution.png')
        
        # 3. 生成报告
        report_paths = {}
        
        if 'excel' in report_formats:
            try:
                if self.excel_generator is None:
                    self.excel_generator = ExcelReportGenerator()
                
                excel_path = output_dir / 'defect_report.xlsx'
                self.excel_generator.generate_defect_report(statistics, str(excel_path))
                report_paths['excel'] = str(excel_path)
            except Exception as e:
                logger.error(f"生成 Excel 报告失败: {e}")
        
        if 'pdf' in report_formats:
            try:
                if self.pdf_generator is None:
                    self.pdf_generator = PDFReportGenerator()
                
                pdf_path = output_dir / 'defect_report.pdf'
                self.pdf_generator.generate_defect_report(
                    statistics,
                    list(figures.values()),
                    str(pdf_path)
                )
                report_paths['pdf'] = str(pdf_path)
            except Exception as e:
                logger.error(f"生成 PDF 报告失败: {e}")
        
        if 'html' in report_formats:
            try:
                html_path = output_dir / 'defect_report.html'
                self.html_generator.generate_defect_report(
                    statistics,
                    chart_paths,
                    str(html_path)
                )
                report_paths['html'] = str(html_path)
            except Exception as e:
                logger.error(f"生成 HTML 报告失败: {e}")
        
        logger.info("报告生成完成！")
        logger.info(f"报告已保存到: {output_dir}")
        
        return {
            'statistics': statistics,
            'figures': figures,
            'report_paths': report_paths
        }
